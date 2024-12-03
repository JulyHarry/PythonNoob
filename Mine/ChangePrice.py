from openpyxl import load_workbook

PRICE1 = 40
PRICE2 = 60
PATH = '/Users/hang/Downloads/报价11-18.xlsx'
SHEET_NAME = 'Sheet1'


class ChangePrice:
    def __init__(self, xs, xn, ys, yn):
        # 加载Excel文件
        self.workbook = load_workbook(PATH)
        # 选择默认的工作表
        self.sheet = self.workbook[SHEET_NAME]
        self.tp = self.coordinate(xs, xn, ys, yn)

    def save(self):
        self.workbook.save(PATH)

    def coordinate(self, xs, xn, ys, yn):
        return [ord(xs) - ord('a') + 1, ord(ys) - ord('a') + 2, int(xn), int(yn) + 1]

    def traversal(self):
        # 遍历工作表中的行和列
        for row in range(self.tp[2], self.tp[3]):
            for col in range(self.tp[0], self.tp[1]):
                cell_value = self.sheet.cell(row=row, column=col).value
                self.sheet.cell(row=row, column=col).value = self.parse(cell_value)
        self.save()

    def parse(self, st):
        st = str(st)
        s = st.split(' ')
        for i in range(len(s)):
            if s[i].isdigit():
                s[i] = self.change(s[i])
        return ' '.join(s)

    def change(self, v):
        v = int(v)
        if v < PRICE1:
            v += 40
        else:
            v += 60
        return str(v)


if __name__ == '__main__':
    cp = ChangePrice('f', '3', 'f', '63')
    # cp = ChangePrice('j', '3', 'j', '57')
    # cp = ChangePrice('k', '27', 'k', '32')
    cp.traversal()
